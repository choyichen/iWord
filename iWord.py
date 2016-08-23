# -*- coding: utf-8 -*-
"""iWord

Get the definitions from various online dictionary.

Note: The display font style & size can be changed via
editing "HKEY_CURRENT_USER\Console\C:_Python27_python.exe"
"""
from sys import exit
from time import sleep
from random import choice
from glob import glob
from difflib import get_close_matches

import urllib
import lxml.html
import os
import os.path

# Set the timeout
TIMEOUT = 20

# Load the word-list
WORDLIST = 'word.txt'
words = {k.lower():v for k,v in [line.rstrip().split('\t', 1) for line in open(WORDLIST)]}

# prepare pydict
pydict = {}
for f in glob(r'lib/*.lib'):
    for line in open(f):
        word, definition = line.strip().split('=', 1)
        pydict[word] = definition

# Load <root.xlsx>
from openpyxl import load_workbook
wb = load_workbook('root.xlsx', True)
def _load_worksheet(wb, sheet):
    D = {}
    for row in wb.get_sheet_by_name(sheet).iter_rows():
        L,j,k = row[0].internal_value, row[1].internal_value, row[2].internal_value
        for i in L.split('/'):
            if i in D:
                raise Exception, "duplicated key '%s' in <%s>" % (L, sheet)
            D[i] = "%s\t%s -- %s" % (i, j, k)
    return D
prefix = _load_worksheet(wb, 'Prefix')
suffix = _load_worksheet(wb, 'Suffix')
root   = _load_worksheet(wb, 'Root')

# Prepare wav dir path & log file
wav_dir = os.path.join(os.getcwd(), 'wav')
if not os.path.isdir(wav_dir):
    os.mkdir(wav_dir)
if not os.path.isfile('wav.txt'):
    open('wav.txt', 'w').write('word\tfile\n')

# -----------------------------------------------------------------------------
# Function definition

def main(word=None):
    """Main flow"""
    global words, pydict
    # prepare the word
    if not word:
        word, value = choice(words.items())  # randomly pick one to show
    elif word.lower() in words:
        value = words[word.lower()]
    else:
        value = ''
    # change the window title
    os.system('title %s' % word + ': ' + value)
    os.system('cls')
    # logging the word
    open('log.txt', 'a').write(word + '\n')
    # print out the results
    header_print(word + ': ' + value)
    speak(word)
    print_prefix(word)
    print
    for i in get_close_matches(word, pydict):
        print i, pydict[i]
    print
    wav_path = mw(word)
    replay(wav_path)
    yahoo(word)
    etymology(word)
    return word, wav_path

def print_prefix(word):
    """Identify the preix, root, and suffix of the given word."""
    word = word.lower()
    for k,v in prefix.iteritems():
        if word.startswith(k):
            print v
    for k,v in root.iteritems():
        if k in word:
            print v
    for k,v in suffix.iteritems():
        if word.endswith(k):
            print v

def replay(path):
    """Play the wav file."""
    import winsound
    if path:
        winsound.PlaySound(path, winsound.SND_FILENAME)
        #winsound.PlaySound(path, winsound.SND_FILENAME | winsound.SND_ASYNC)

def speak(sentence):
    """Use Windows TTS to speak a sentence."""
    import win32com.client
    speaker = win32com.client.Dispatch("SAPI.SpVoice")
    speaker.Speak(sentence)

def header_print(s):
    """Print the string s as a header."""
    print s
    print '-' * len(s)

def safe_print(e):
    """Print the text content of an element."""
    from string import printable
    if e is not None:
        #print e.text_content()
        s = e.text_content()
        print "".join(i for i in s if i in printable)

def mw(word):
    """Parse the results from m-w Dictionary & 
    download the pronunciation wav file to a temporary place, and return the path.
    """
    url = "http://www.merriam-webster.com/dictionary/{}".format(word)
    try:
        tree = lxml.html.parse(url)
    except IOError:
        print "Word not found in m-w.\n"
        return
    item = tree.find('.//div[@class="headword"]/h2').text_content()
    if item:
        header_print(item)
        # if only one definition
        for i in tree.findall('.//div[@class="scnt"]/span'):
            safe_print(i)
#        for i in tree.findall('.//div[@class="sense-block-one"]'):
#            for j in i.findall('./div[@class="scnt"]/span'):
#                safe_print(j)
#        # multiple definitions
#        for i in tree.findall('.//div[@class="sblk"]'):  # def
#            n = i.xpath('./div[@class="snum"]')[0].text  # no.
#            for j in i.findall('./div[@class="scnt"]/span'):
#                print n,
#                safe_print(j)
        print
        safe_print(tree.find('.//div[@class="syn-para"]'))
        safe_print(tree.find('.//div[@class="ant-para"]'))
        print
        # Get the pronunciation wav URL & file path
        au_btn = tree.find('.//input[@class="au"]')
        if au_btn is not None:
            au = au_btn.attrib['onclick'].split("'")[1] + '.wav'
            path = os.path.join(wav_dir, au)
            # Search pre-downloaded wav files. If the wav not existed, download it.
            if not os.path.exists(path):
                url = "http://media.merriam-webster.com/soundc11/%s/%s" % (au[0], au)
                urllib.urlretrieve(url, path)
                # write the word-wav mapping to log file
                with open('wav.txt', 'a') as fout:
                    fout.write("%s\t%s\n" % (word, au))
            return path

def yahoo(word):
    """Parsing the results from Yahoo! Dictionary."""
    url = "http://tw.dictionary.yahoo.com/dictionary?p={}".format(word)
    tree = lxml.html.parse(url)
    item = tree.find('.//div[@class="theme clr"]')
    if item is not None:
        header_print(item.text_content())
        for i in tree.findall('.//div[@class="def clr nobr"]'):
            print i.xpath("div")[0].text  # caption
            for n, j in enumerate(i.findall('.//li')):
                print '  ' + "{:<2}".format(n+1) + j.xpath("div/p")[0].text  # interpret
                for k in j.xpath("p"):
                    Eng, Chn = k.text_content().split('\n')  # examples
                    print '  ' + Eng
                    print '  ' + Chn
                    speak(Eng)
                print
        #safe_print(tree.find('.//div[@class="synonyms"]'))
        #safe_print(tree.find('.//div[@class="variation"]'))
        #print
    else:
        print 'Word not found in Yahoo!\n'

def etymology(word):
    """Parsing the results from Online Etymology Dictionary."""
    url = "http://www.etymonline.com/index.php?allowed_in_frame=0&search={}&searchmode=none".format(word)
    tree = lxml.html.parse(url)
    item = tree.find('.//dt[@class="highlight"]')
    if item is not None:
        item = item.text_content().strip()
        header_print(item)
        safe_print(tree.find('.//dd[@class="highlight"]'))
#        try:
#            print tree.find('.//dd[@class="highlight"]').text_content()
#        except UnicodeEncodeError:
#            print 'Unpritable chracters found.'
        print

def show_web_page(word):
    import webbrowser
    webbrowser.open('http://www.csie.ntu.edu.tw/~r96945018/dict.php?word=%s' % word)

def help():
    print 'Avaliable commands:'
    print 'n or <enter> --> next word'
    print 'r --> replay the pronunciation'
    print 'y --> repeat the example sentences'
    print 'w --> show the web page'
    print 'd --> delete the word from the word-list'
    print 's --> save the word to the word-list or give a new definition'
    print 'q --> quit'
    print 'any single chracter --> help'
    print 'any word --> search the word'

def save(word):
    """Save the word to current word list."""
    global words
    if word in words:
        # the word already has a definition in the word-list
        print "Current definition:", words[word]
        chn = raw_input("Enter a new definition (q to leave): ").strip()
        if chn != 'q':
            words[word] = chn
            lines = [L for L in open(WORDLIST) if not L.startswith(word + '\t')]  # remove old records
            lines.append("{}\t{}\n".format(word, chn))  # add a new record
            open(WORDLIST, 'w').writelines(lines)
    else:
        # the word is not in the word-list
        chn = raw_input("Enter a short definition: ").strip()
        words[word] = chn
        open(WORDLIST, 'a').write("{}\t{}\n".format(word, chn))

def delete(word):
    """Remove the word from the word-list."""
    global words
    del words[word]
    lines = [L for L in open(WORDLIST) if not L.lower().startswith(word + '\t')]  # remove old records
    open(WORDLIST, 'w').writelines(lines)
    print word, 'deleted'

def listwords():
    global words
    for k,v in words.iteritems():
        print k, '\t', v
    print '-' * 30
    print 'Total {} words in <{}>.'.format(len(words), WORDLIST)

if __name__ == "__main__":
    cmd = 'h'  # default command is n
    while True:
        if cmd == 'n':  # next-word
            word, path = main()
            try:
                print "Hit <ctrl-c> for new commands...  ",
                # count down till timeout
                #sleep(TIMEOUT)  # wait for command
                for i in range(TIMEOUT, 0, -1):
                    print "\b\b\b{:2}".format(i),
                    sleep(1) # wait for command
                print '\b'
                continue  # automatically proceed to next word
            except KeyboardInterrupt:
                pass  # interupted by users
        elif cmd == 'r':
            replay(path)
        elif cmd == 'w':
            show_web_page(word)
        elif cmd == 'd':
            delete(word)
        elif cmd == 's':
            save(word)
        elif cmd == 'l':
            listwords()
        elif cmd == 'y':
            yahoo(word)
        elif cmd == 'q':
            exit()
        elif len(cmd) == 1:
            help()
        else:
            # query the the word as user inputed
            word, path = main(cmd)
        # wait for user's command
        cmd = raw_input('>>> ').strip()
